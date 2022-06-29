from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views import View
from casa.forms import *
from casa.models import *

from django.views.generic.edit import DeleteView, CreateView, UpdateView
from django.views.generic import ListView, TemplateView
from openpyxl import Workbook, Workbook

import os
from django.conf import settings
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders




# Create your views here.

#PROPIETARIOS

class RegistrarPropietario(CreateView):
    model = Propietario
    form_class = FormularioPropietario
    template_name = "registrarpropietario.html"    
    success_url = reverse_lazy('propietarios')

class EliminarPropietario(DeleteView):
    model = Propietario
    template_name = "propietario_eliminar.html"
    success_url = reverse_lazy('propietarios')      


class ListarPropietarios(ListView):
    model = Propietario                             
    template_name = "propietarios.html"

class ActualizarPropietario(UpdateView):            
    model = Propietario
    template_name = "editarpropietario.html"
    form_class = FormularioPropietario
    success_url = reverse_lazy('propietarios')


#INQUILINOS

class RegistrarInquilino(CreateView):
    model = Inquilino
    form_class = FormularioInquilino
    template_name = "registrarinquilino.html"     
    success_url = reverse_lazy('inquilinos')


class ActualizarInquilino(UpdateView):            
    model = Inquilino
    template_name = "editarinquilino.html"
    form_class = FormularioInquilino
    success_url = reverse_lazy('inquilinos')


class EliminarInquilino(DeleteView):
    model = Inquilino
    template_name = "inquilino_eliminar.html"
    success_url = reverse_lazy('inquilinos')

class ListarInquilinos(ListView):
    model = Inquilino
    template_name = "inquilinos.html"

#CASA

class ListarCasas(ListView):
    model = Casa
    template_name = "listado_casas.html"

class ActualizarCasa(UpdateView):            
    model = Casa
    template_name = "editarcasa.html"
    form_class = FormularioCasa
    success_url = reverse_lazy('listado_casas')


class FormularioCasaView(HttpRequest):

    def editar_casa(request, id_casa):
        casa = Casa.objects.filter(id=id_casa).first()
        form = FormularioCasa(instance = casa)
        return render(request, 'editarcasa.html', {'form':form, 'casa':casa})

    def actualizar_casa(request, id_casa):
        casa = Casa.objects.get(pk=id_casa)
        form = FormularioCasa(request.POST, instance=casa)
        if form.is_valid():
            form.save()
        casas = Casa.objects.all()
        return render(request, 'listado_casas.html', {'casas': casas})

#EXPORTAR XLS (hay que instalar openpyxl)

class ExportarPropietarios(TemplateView):
    
    def get(self, request, *args):
        propietarios = Propietario.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Propietarios'
        ws['B1'] = 'REPORTE DE PROPIETARIOS'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'Nombre'
        ws['C3'] = 'Apellido'
        ws['D3'] = 'Telefono'
        ws['E3'] = 'Email'
        ws['F3'] = 'CBU'
        contador = 4        #a partir de donde va a empezar la tabla
        for propietario in propietarios:
            ws.cell(row = contador, column = 2).value = propietario.nombre
            ws.cell(row = contador, column = 3).value = propietario.apellido
            ws.cell(row = contador, column = 4).value = propietario.telefono
            ws.cell(row = contador, column = 5).value = propietario.email
            ws.cell(row = contador, column = 6).value = propietario.CBU
            contador += 1   #a medida q encuentra aumenta una fila

        nombre_archivo = 'reporte_propietarios.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class ExportarInquilinos(TemplateView):
    
    def get(self, request, *args):
        inquilinos = Inquilino.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inquilinos'
        ws['B1'] = 'REPORTE DE PROPIETARIOS'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'Nombre'
        ws['C3'] = 'Apellido'
        ws['D3'] = 'Telefono'
        ws['E3'] = 'Email'
        ws['F3'] = 'CBU'
        contador = 4        #a partir de donde va a empezar la tabla
        for inquilino in inquilinos:
            ws.cell(row = contador, column = 2).value = inquilino.nombre
            ws.cell(row = contador, column = 3).value = inquilino.apellido
            ws.cell(row = contador, column = 4).value = inquilino.telefono
            ws.cell(row = contador, column = 5).value = inquilino.email
            ws.cell(row = contador, column = 6).value = inquilino.CBU

            contador += 1   #a medida q encuentra aumenta una fila

        nombre_archivo = 'reporte_inquilinos.xlsx'
        response = HttpResponse(content_type = 'application/ms-excel')
        content = 'attachment; filename = {0}'.format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

#Exportar PDFs (Hay que instalar xhtml2pdf)

class ExportarPropietariosPDF(View):
    def get(self, request, *args, **kwargs):
        return HttpResponse('Hello, World!')





